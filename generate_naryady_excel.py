from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference


PERSONNEL_ROWS = 300
DUTY_ROWS = 500


def setup_header(row, titles, bold=True):
    for col_idx, title in enumerate(titles, start=1):
        cell = row.parent.cell(row=row.row, column=col_idx, value=title)
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_reference_sheet(wb):
    ws = wb.create_sheet("Довідник")
    ws["A1"] = "Звання"
    ws["A1"].font = Font(bold=True)

    ranks = [
        "Солдат",
        "Старший солдат",
        "Молодший сержант",
        "Сержант",
        "Старший сержант",
        "Прапорщик",
        "Лейтенант",
        "Старший лейтенант",
        "Капітан",
        "Майор",
    ]

    for i, rank in enumerate(ranks, start=2):
        ws[f"A{i}"] = rank

    set_column_widths(ws, {"A": 25})
    ws.auto_filter.ref = "A1:A11"
    return ws


def create_personnel_sheet(wb, ref_sheet):
    ws = wb.create_sheet("Особовий_склад")
    headers = ["№", "Посада", "Звання", "ФІО"]
    setup_header(ws["A1"], headers)

    # Автонумерация для наглядности (можно перезаписать руками)
    for row in range(2, PERSONNEL_ROWS + 2):
        ws[f"A{row}"] = row - 1

    # Ширина колонок
    set_column_widths(
        ws,
        {
            "A": 6,
            "B": 25,
            "C": 20,
            "D": 30,
        },
    )

    # Фильтры
    ws.auto_filter.ref = f"A1:D{PERSONNEL_ROWS + 1}"

    # Выпадающий список званий
    dv_rank = DataValidation(
        type="list",
        formula1="=Довідник!$A$2:$A$11",
        allow_blank=True,
        showDropDown=True,
        error="Оберіть звання зі списку.",
        errorTitle="Невірне значення",
    )
    ws.add_data_validation(dv_rank)
    dv_rank.add(f"C2:C{PERSONNEL_ROWS + 1}")

    return ws


def create_duty_sheet(wb):
    ws = wb.create_sheet("Наряды")
    headers = ["№", "Посада", "Звание", "ФІО", "Дата заступления"]
    setup_header(ws["A1"], headers)

    # Автоматическая нумерация
    for row in range(2, DUTY_ROWS + 2):
        ws[f"A{row}"] = f"=ROW()-1"

    # Ширина колонок
    set_column_widths(
        ws,
        {
            "A": 6,
            "B": 25,
            "C": 20,
            "D": 30,
            "E": 18,
        },
    )

    # Фильтры
    ws.auto_filter.ref = f"A1:E{DUTY_ROWS + 1}"

    # Выпадающий список званий
    dv_rank = DataValidation(
        type="list",
        formula1="=Довідник!$A$2:$A$11",
        allow_blank=True,
        showDropDown=True,
        error="Оберіть звання зі списку.",
        errorTitle="Невірне значення",
    )
    ws.add_data_validation(dv_rank)
    dv_rank.add(f"C2:C{DUTY_ROWS + 1}")

    # Валидация даты (просто тип дата)
    dv_date = DataValidation(type="date", allow_blank=True)
    ws.add_data_validation(dv_date)
    dv_date.add(f"E2:E{DUTY_ROWS + 1}")

    return ws


def create_release_sheet(wb):
    ws = wb.create_sheet("Звільнення")
    headers = ["№", "Посада", "Звание", "ФІО", "Дата звільнення"]
    setup_header(ws["A1"], headers)

    # Нумерация по формуле
    for row in range(2, PERSONNEL_ROWS + 2):
        ws[f"A{row}"] = f"=ROW()-1"

    set_column_widths(
        ws,
        {
            "A": 6,
            "B": 25,
            "C": 20,
            "D": 30,
            "E": 18,
        },
    )

    ws.auto_filter.ref = f"A1:E{PERSONNEL_ROWS + 1}"

    # Выпадающий список званий
    dv_rank = DataValidation(
        type="list",
        formula1="=Довідник!$A$2:$A$11",
        allow_blank=True,
        showDropDown=True,
        error="Оберіть звання зі списку.",
        errorTitle="Невірне значення",
    )
    ws.add_data_validation(dv_rank)
    dv_rank.add(f"C2:C{PERSONNEL_ROWS + 1}")

    # Валидация даты
    dv_date = DataValidation(type="date", allow_blank=True)
    ws.add_data_validation(dv_date)
    dv_date.add(f"E2:E{PERSONNEL_ROWS + 1}")

    return ws


def create_statistics_sheet(wb):
    ws = wb.create_sheet("Статистика")
    headers = ["ФІО", "Кількість нарядів"]
    setup_header(ws["A1"], headers)

    # ФИО подтягиваем из листа "Особовий_склад"
    for row in range(2, PERSONNEL_ROWS + 2):
        ws[f"A{row}"] = f"=IF(Особовий_склад!D{row}=\"\",\"\",Особовий_склад!D{row})"
        ws[f"B{row}"] = (
            f"=IF(A{row}=\"\",\"\",COUNTIF(Наряды!$D$2:$D${DUTY_ROWS + 1},A{row}))"
        )

    set_column_widths(ws, {"A": 30, "B": 18})
    ws.auto_filter.ref = f"A1:B{PERSONNEL_ROWS + 1}"

    # Условное форматирование для максимума и минимума
    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    max_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Максимальное количество нарядов
    ws.conditional_formatting.add(
        f"B2:B{PERSONNEL_ROWS + 1}",
        FormulaRule(
            formula=[f"B2=MAX($B$2:$B${PERSONNEL_ROWS + 1})"],
            stopIfTrue=False,
            fill=max_fill,
        ),
    )

    # Минимальное (больше 0, чтобы не подсвечивать всех с нулём)
    ws.conditional_formatting.add(
        f"B2:B{PERSONNEL_ROWS + 1}",
        FormulaRule(
            formula=[
                f'AND(B2>0,B2=MIN(IF($B$2:$B${PERSONNEL_ROWS + 1}>0,$B$2:$B${PERSONNEL_ROWS + 1})))'
            ],
            stopIfTrue=False,
            fill=min_fill,
        ),
    )

    return ws


def create_queue_sheet(wb):
    ws = wb.create_sheet("Черга")

    ws["A1"] = "Наступний в наряд"
    ws["A1"].font = Font(bold=True)
    ws["A3"] = "Останній в наряді"
    ws["A3"].font = Font(bold=True)

    # Последний человек в наряде (по ФИО)
    ws["B3"] = (
        "=IFERROR(LOOKUP(2,1/(Наряды!$D$2:$D$" f"{DUTY_ROWS + 1}" ' <>""),Наряды!$D$2:$D$'
        f"{DUTY_ROWS + 1}" "),\"\")"
    )

    # Следующий в наряд:
    # Используем наименьшее количество нарядов из "Статистика",
    # исключая пустые ФИО и последнего, кто был в наряде.
    ws["B1"] = (
        "=IFERROR("
        "INDEX(Статистика!$A$2:$A$" f"{PERSONNEL_ROWS + 1}" ","
        "MATCH("
        "MINIFS(Статистика!$B$2:$B$" f"{PERSONNEL_ROWS + 1}"
        ",Статистика!$A$2:$A$" f"{PERSONNEL_ROWS + 1}" ',\"<>\"'
        ",Статистика!$A$2:$A$" f"{PERSONNEL_ROWS + 1}" ",\"<>\"&B3"
        "),"
        "Статистика!$B$2:$B$" f"{PERSONNEL_ROWS + 1}" ",0)"
        "),\"\")"
    )

    set_column_widths(ws, {"A": 25, "B": 35})

    return ws


def create_charts_sheet(wb, stats_sheet):
    ws = wb.create_sheet("Графики")

    ws["A1"] = "Діаграма кількості нарядів по людям"
    ws["A1"].font = Font(bold=True)

    # Диаграмма по диапазону Статистика!A2:B{PERSONNEL_ROWS+1}
    data = Reference(
        stats_sheet,
        min_col=2,
        max_col=2,
        min_row=1,
        max_row=PERSONNEL_ROWS + 1,
    )
    cats = Reference(
        stats_sheet,
        min_col=1,
        max_col=1,
        min_row=2,
        max_row=PERSONNEL_ROWS + 1,
    )

    chart = BarChart()
    chart.title = "Кількість нарядів"
    chart.y_axis.title = "Наряди"
    chart.x_axis.title = "Особовий склад"
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 30
    chart.height = 15

    ws.add_chart(chart, "A3")

    return ws


def protect_sheets(wb):
    # Простая защита формул от случайного изменения
    # (пароль очень простой, чтобы можно было легко снять)
    password = "1234"

    for ws in wb.worksheets:
        # Блокируем только ячейки с формулами
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    # В новых версиях openpyxl стили неизменяемые, поэтому копируем protection
                    cell.protection = cell.protection.copy(locked=True)
        ws.protection.set_password(password)
        ws.protection.enable()


def main():
    wb = Workbook()
    # По умолчанию создаётся один лист, удалим его
    default_sheet = wb.active
    wb.remove(default_sheet)

    ref_sheet = create_reference_sheet(wb)
    create_personnel_sheet(wb, ref_sheet)
    create_duty_sheet(wb)
    create_release_sheet(wb)
    stats_sheet = create_statistics_sheet(wb)
    create_queue_sheet(wb)
    create_charts_sheet(wb, stats_sheet)
    protect_sheets(wb)

    output_name = "Учет_военных_нарядов.xlsx"
    wb.save(output_name)
    print(f"Файл '{output_name}' успешно создан.")


if __name__ == "__main__":
    main()

